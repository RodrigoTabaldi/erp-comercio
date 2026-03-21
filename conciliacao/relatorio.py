import openpyxl                                    # biblioteca para criar arquivos Excel
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter       # converte número de coluna em letra (1 → A)
from datetime import datetime                       # para colocar a data no nome do arquivo
import os                                           # para criar pastas se não existirem


# ---------------------------------------------------------
# CORES — definidas aqui em cima para facilitar mudanças
# ---------------------------------------------------------

# Cabeçalhos das abas
COR_CABECALHO_VERDE   = 'FF1E7E34'  # verde escuro  → aba conciliados
COR_CABECALHO_AMARELO = 'FF856404'  # amarelo escuro → aba só no banco
COR_CABECALHO_VERMELHO= 'FF842029'  # vermelho escuro→ aba só no interno

# Linhas de dados
COR_LINHA_VERDE_CLARO  = 'FFD1E7DD'  # verde claro   → conciliados
COR_LINHA_AMARELO_CLARO= 'FFFFF3CD'  # amarelo claro → só no banco
COR_LINHA_VERMELHO_CLARO='FFF8D7DA'  # vermelho claro→ só no interno
COR_LINHA_BRANCO       = 'FFFFFFFF'  # branco        → linhas alternadas

# Aba de resumo
COR_TITULO_RESUMO = 'FF0D6EFD'  # azul
COR_FUNDO_RESUMO  = 'FFE7F1FF'  # azul claro

# Fonte branca para cabeçalhos escuros
FONTE_BRANCA = Font(color='FFFFFFFF', bold=True, size=11)
FONTE_NORMAL = Font(size=11)
FONTE_TITULO = Font(size=14, bold=True)
FONTE_NEGRITO= Font(size=11, bold=True)


def _criar_borda():
    """Cria uma borda fina para as células."""
    lado = Side(style='thin', color='FFCCCCCC')
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _formatar_cabecalho(ws, colunas, cor_fundo):
    """
    Escreve e formata a linha de cabeçalho de uma aba.

    Parâmetros:
        ws      : worksheet (a aba do Excel)
        colunas : lista com os nomes das colunas
        cor_fundo: código hexadecimal da cor de fundo
    """
    preenchimento = PatternFill(fill_type='solid', fgColor=cor_fundo)
    alinhamento   = Alignment(horizontal='center', vertical='center')
    borda         = _criar_borda()

    for col_num, nome in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_num, value=nome)
        celula.fill      = preenchimento
        celula.font      = FONTE_BRANCA
        celula.alignment = alinhamento
        celula.border    = borda

    # Altura da linha de cabeçalho
    ws.row_dimensions[1].height = 25


def _escrever_dados(ws, df, cor_principal, col_data=None):
    """
    Escreve os dados do DataFrame na aba, com linhas alternadas.

    Parâmetros:
        ws           : worksheet (aba do Excel)
        df           : DataFrame com os dados
        cor_principal: cor das linhas ímpares
        col_data     : índice da coluna de data (para formatar como data)
    """
    preench_principal = PatternFill(fill_type='solid', fgColor=cor_principal)
    preench_branco    = PatternFill(fill_type='solid', fgColor=COR_LINHA_BRANCO)
    borda             = _criar_borda()

    for row_num, (_, linha) in enumerate(df.iterrows(), start=2):
        # Alterna cores: linha par = branco, linha ímpar = cor principal
        preenchimento = preench_principal if row_num % 2 == 0 else preench_branco

        for col_num, valor in enumerate(linha, start=1):
            celula = ws.cell(row=row_num, column=col_num, value=valor)
            celula.fill   = preenchimento
            celula.font   = FONTE_NORMAL
            celula.border = borda

            # Formata coluna de data
            if col_data and col_num == col_data:
                celula.number_format = 'DD/MM/YYYY'
                celula.alignment = Alignment(horizontal='center')

            # Formata coluna de valor (moeda)
            elif isinstance(valor, float):
                celula.number_format = 'R$ #,##0.00'
                celula.alignment = Alignment(horizontal='right')

                # Valor negativo fica vermelho, positivo fica verde
                if valor < 0:
                    celula.font = Font(size=11, color='FF842029')
                else:
                    celula.font = Font(size=11, color='FF1E7E34')

            else:
                celula.alignment = Alignment(horizontal='left')

        ws.row_dimensions[row_num].height = 20


def _ajustar_largura_colunas(ws):
    """Ajusta a largura de cada coluna para caber o conteúdo."""
    for col in ws.columns:
        largura_max = 0
        for celula in col:
            if celula.value:
                largura_max = max(largura_max, len(str(celula.value)))
        # Adiciona uma margem e limita a largura máxima
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(largura_max + 4, 40)


def _criar_aba_resumo(wb, resultado):
    """
    Cria a primeira aba com o resumo geral da conciliação.

    Parâmetros:
        wb       : workbook (arquivo Excel)
        resultado: dicionário retornado pelo conciliador.py
    """
    ws = wb.create_sheet('Resumo', 0)  # 0 = primeira posição

    conciliados   = resultado['conciliados']
    so_no_banco   = resultado['so_no_banco']
    so_no_interno = resultado['so_no_interno']

    total = len(conciliados) + len(so_no_banco) + len(so_no_interno)
    pct   = (len(conciliados) / total * 100) if total > 0 else 0

    # --- Título ---
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value     = 'RELATÓRIO DE CONCILIAÇÃO BANCÁRIA'
    titulo.font      = FONTE_TITULO
    titulo.fill      = PatternFill(fill_type='solid', fgColor=COR_TITULO_RESUMO)
    titulo.font      = Font(size=14, bold=True, color='FFFFFFFF')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # --- Data de geração ---
    ws.merge_cells('A2:D2')
    data_cel = ws['A2']
    data_cel.value     = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
    data_cel.font      = Font(size=10, italic=True, color='FF666666')
    data_cel.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10  # espaço em branco

    # --- Cards de resumo ---
    dados_resumo = [
        ('DESCRIÇÃO',             'QUANTIDADE', 'VALOR TOTAL',   'STATUS'),
        ('Transações conciliadas', len(conciliados),
            conciliados['valor'].sum() if len(conciliados) > 0 else 0,   '✅ OK'),
        ('Só no banco',            len(so_no_banco),
            so_no_banco['valor'].sum() if len(so_no_banco) > 0 else 0,   '⚠️ Atenção'),
        ('Só no interno',          len(so_no_interno),
            so_no_interno['valor'].sum() if len(so_no_interno) > 0 else 0,'❌ Divergência'),
        ('TOTAL GERAL',            total, '', ''),
    ]

    cores_resumo = [
        COR_CABECALHO_VERDE,    # cabeçalho
        COR_LINHA_VERDE_CLARO,  # conciliados
        COR_LINHA_AMARELO_CLARO,# só no banco
        COR_LINHA_VERMELHO_CLARO,# só no interno
        'FFE2E2E2',             # total
    ]

    for i, (linha, cor) in enumerate(zip(dados_resumo, cores_resumo), start=4):
        preenchimento = PatternFill(fill_type='solid', fgColor=cor)
        borda         = _criar_borda()

        for j, valor in enumerate(linha, start=1):
            celula = ws.cell(row=i, column=j, value=valor)
            celula.fill   = preenchimento
            celula.border = borda
            celula.alignment = Alignment(horizontal='center', vertical='center')

            if i == 4:  # linha de cabeçalho do resumo
                celula.font = FONTE_BRANCA
            elif j == 3 and isinstance(valor, float):  # coluna de valor
                celula.number_format = 'R$ #,##0.00'
                celula.font = FONTE_NEGRITO
            elif i == 8:  # linha total
                celula.font = FONTE_NEGRITO
            else:
                celula.font = FONTE_NORMAL

        ws.row_dimensions[i].height = 22

    # --- Taxa de conciliação ---
    ws.row_dimensions[9].height = 10
    ws.merge_cells('A10:D10')
    taxa = ws['A10']
    taxa.value     = f'Taxa de conciliação: {pct:.1f}%'
    taxa.font      = Font(size=12, bold=True, color='FF0D6EFD')
    taxa.alignment = Alignment(horizontal='center')
    taxa.fill      = PatternFill(fill_type='solid', fgColor=COR_FUNDO_RESUMO)
    ws.row_dimensions[10].height = 25

    # Ajusta largura das colunas do resumo
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def gerar_relatorio(resultado, caminho_saida=None):
    """
    Gera o relatório Excel completo com 4 abas:
        1. Resumo        → visão geral da conciliação
        2. Conciliados   → transações que batem nos dois lados
        3. Só no banco   → transações não lançadas internamente
        4. Só no interno → lançamentos que não aparecem no banco

    Parâmetros:
        resultado     : dicionário retornado pelo conciliador.py
        caminho_saida : onde salvar o Excel (opcional)

    Retorna:
        O caminho onde o arquivo foi salvo.
    """

    # Define o nome do arquivo com a data atual
    if caminho_saida is None:
        os.makedirs('relatorios', exist_ok=True)
        data_hoje = datetime.now().strftime('%Y%m%d_%H%M%S')
        caminho_saida = f'relatorios/conciliacao_{data_hoje}.xlsx'

    # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba vazia padrão

    # ---------------------------------------------------------
    # ABA 1 — Resumo
    # ---------------------------------------------------------
    _criar_aba_resumo(wb, resultado)

    # ---------------------------------------------------------
    # ABA 2 — Conciliados ✅
    # ---------------------------------------------------------
    ws_conc = wb.create_sheet('✅ Conciliados')
    df_conc = resultado['conciliados']

    if len(df_conc) > 0:
        colunas = [c.replace('_banco', ' (banco)').replace('_interno', ' (interno)').upper()
                   for c in df_conc.columns]
        _formatar_cabecalho(ws_conc, colunas, COR_CABECALHO_VERDE)
        _escrever_dados(ws_conc, df_conc, COR_LINHA_VERDE_CLARO, col_data=1)
        _ajustar_largura_colunas(ws_conc)
    else:
        ws_conc['A1'] = 'Nenhuma transação conciliada encontrada.'

    # ---------------------------------------------------------
    # ABA 3 — Só no banco ⚠️
    # ---------------------------------------------------------
    ws_banco = wb.create_sheet('⚠️ Só no banco')
    df_banco = resultado['so_no_banco']

    if len(df_banco) > 0:
        colunas = [c.replace('_banco', ' (banco)').replace('_interno', ' (interno)').upper()
                   for c in df_banco.columns]
        _formatar_cabecalho(ws_banco, colunas, COR_CABECALHO_AMARELO)
        _escrever_dados(ws_banco, df_banco, COR_LINHA_AMARELO_CLARO, col_data=1)
        _ajustar_largura_colunas(ws_banco)
    else:
        ws_banco['A1'] = 'Nenhuma divergência encontrada no banco.'

    # ---------------------------------------------------------
    # ABA 4 — Só no interno ❌
    # ---------------------------------------------------------
    ws_int = wb.create_sheet('❌ Só no interno')
    df_int = resultado['so_no_interno']

    if len(df_int) > 0:
        colunas = [c.replace('_banco', ' (banco)').replace('_interno', ' (interno)').upper()
                   for c in df_int.columns]
        _formatar_cabecalho(ws_int, colunas, COR_CABECALHO_VERMELHO)
        _escrever_dados(ws_int, df_int, COR_LINHA_VERMELHO_CLARO, col_data=1)
        _ajustar_largura_colunas(ws_int)
    else:
        ws_int['A1'] = 'Nenhuma divergência encontrada nos lançamentos internos.'

    # ---------------------------------------------------------
    # Salva o arquivo
    # ---------------------------------------------------------
    wb.save(caminho_saida)
    print(f"\n✅ Relatório gerado com sucesso!")
    print(f"📁 Salvo em: {caminho_saida}")

    return caminho_saida


# ---------------------------------------------------------
# TESTE — usa os mesmos dados do conciliador.py
# ---------------------------------------------------------
if __name__ == '__main__':
    import pandas as pd

    # Simula o resultado que viria do conciliador.py
    resultado = {
        'conciliados': pd.DataFrame([
            {'data': '2026-03-01', 'descricao_banco': 'SALARIO',               'valor':  5000.00, 'descricao_interno': 'SALARIO'},
            {'data': '2026-03-02', 'descricao_banco': 'SUPERMERCADO',          'valor':  -350.75, 'descricao_interno': 'SUPERMERCADO'},
            {'data': '2026-03-05', 'descricao_banco': 'CONTA DE LUZ',          'valor':  -189.90, 'descricao_interno': 'CONTA DE LUZ'},
            {'data': '2026-03-10', 'descricao_banco': 'TRANSFERENCIA RECEBIDA','valor':  1200.00, 'descricao_interno': 'TRANSFERENCIA RECEBIDA'},
            {'data': '2026-03-15', 'descricao_banco': 'FARMACIA',              'valor':   -87.50, 'descricao_interno': 'FARMACIA'},
        ]),
        'so_no_banco': pd.DataFrame([
            {'data': '2026-03-18', 'descricao_banco': 'TAXA BANCARIA', 'valor': -25.00, 'descricao_interno': None},
        ]),
        'so_no_interno': pd.DataFrame([
            {'data': '2026-03-20', 'descricao_banco': None, 'valor': -500.00, 'descricao_interno': 'FORNECEDOR XYZ'},
        ]),
    }

    print("=== Gerando relatório de teste ===")
    caminho = gerar_relatorio(resultado)
    print(f"\nAbra o arquivo para ver o resultado: {caminho}")
