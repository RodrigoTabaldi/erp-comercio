import flet as ft
from database import buscar_todos, resumo_financeiro
from datetime import datetime
import os


def snack(page, msg, cor="#1e4a2e"):
    page.snack_bar = ft.SnackBar(ft.Text(msg), bgcolor=cor)
    page.snack_bar.open = True
    page.update()


def tela(page, COR_CARD, COR_BORDA, COR_AZUL, COR_TEXTO, COR_SUBTEXTO):

    area_resultado = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)

    def relatorio_clientes(e):
        area_resultado.controls.clear()
        clientes = buscar_todos("SELECT * FROM clientes WHERE ativo = 1 ORDER BY nome")
        area_resultado.controls.append(
            ft.Text(f"Clientes cadastrados ({len(clientes)})", size=16, weight=ft.FontWeight.BOLD, color=COR_TEXTO)
        )
        for c in clientes:
            area_resultado.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Text(c['nome'], size=13, weight=ft.FontWeight.W_500, color=COR_TEXTO),
                            ft.Text(f"{c['telefone'] or '-'} • {c['email'] or '-'} • {c['cidade'] or '-'}", size=12, color=COR_SUBTEXTO),
                        ], spacing=2, expand=True),
                    ]),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
                    border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=10)
                )
            )
        page.update()

    def relatorio_estoque(e):
        area_resultado.controls.clear()
        produtos = buscar_todos("SELECT * FROM produtos WHERE ativo = 1 ORDER BY nome")
        baixo    = [p for p in produtos if p['estoque_atual'] <= p['estoque_minimo']]
        normal   = [p for p in produtos if p['estoque_atual'] > p['estoque_minimo']]

        area_resultado.controls.append(
            ft.Text(f"Relatório de Estoque — {len(produtos)} produtos", size=16, weight=ft.FontWeight.BOLD, color=COR_TEXTO)
        )
        area_resultado.controls.append(
            ft.Row([
                ft.Container(content=ft.Column([ft.Text("Total", size=12, color=COR_SUBTEXTO), ft.Text(str(len(produtos)), size=24, weight=ft.FontWeight.BOLD, color=COR_AZUL)], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
                ft.Container(content=ft.Column([ft.Text("Estoque baixo", size=12, color=COR_SUBTEXTO), ft.Text(str(len(baixo)), size=24, weight=ft.FontWeight.BOLD, color="#f44747")], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
                ft.Container(content=ft.Column([ft.Text("Normal", size=12, color=COR_SUBTEXTO), ft.Text(str(len(normal)), size=24, weight=ft.FontWeight.BOLD, color="#4ec9b0")], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
            ], spacing=12)
        )
        if baixo:
            area_resultado.controls.append(ft.Text("⚠️ Estoque baixo", size=14, weight=ft.FontWeight.BOLD, color="#dcdcaa"))
            for p in baixo:
                area_resultado.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Text(p['nome'], size=13, color=COR_TEXTO, expand=True),
                            ft.Text(f"Atual: {p['estoque_atual']}", size=12, color="#f44747"),
                            ft.Text(f"  Mín: {p['estoque_minimo']}", size=12, color=COR_SUBTEXTO),
                            ft.Text(f"  R$ {p['preco_venda']:.2f}", size=12, color=COR_AZUL),
                        ]),
                        bgcolor=COR_CARD, border=ft.border.all(1, "#f44747"),
                        border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=10)
                    )
                )
        area_resultado.controls.append(ft.Text("✅ Estoque normal", size=14, weight=ft.FontWeight.BOLD, color="#4ec9b0"))
        for p in normal:
            area_resultado.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Text(p['nome'], size=13, color=COR_TEXTO, expand=True),
                        ft.Text(f"Atual: {p['estoque_atual']}", size=12, color="#4ec9b0"),
                        ft.Text(f"  Mín: {p['estoque_minimo']}", size=12, color=COR_SUBTEXTO),
                        ft.Text(f"  R$ {p['preco_venda']:.2f}", size=12, color=COR_AZUL),
                    ]),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
                    border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=10)
                )
            )
        page.update()

    def relatorio_financeiro(e):
        area_resultado.controls.clear()
        resumo   = resumo_financeiro()
        receitas = buscar_todos("SELECT * FROM financeiro WHERE tipo='receita' ORDER BY vencimento DESC")
        despesas = buscar_todos("SELECT * FROM financeiro WHERE tipo='despesa' ORDER BY vencimento DESC")

        area_resultado.controls.append(ft.Text("Relatório Financeiro", size=16, weight=ft.FontWeight.BOLD, color=COR_TEXTO))
        area_resultado.controls.append(
            ft.Row([
                ft.Container(content=ft.Column([ft.Text("Receitas pagas", size=12, color=COR_SUBTEXTO), ft.Text(f"R$ {resumo['receitas']:,.2f}", size=20, weight=ft.FontWeight.BOLD, color="#4ec9b0")], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
                ft.Container(content=ft.Column([ft.Text("Despesas pagas", size=12, color=COR_SUBTEXTO), ft.Text(f"R$ {resumo['despesas']:,.2f}", size=20, weight=ft.FontWeight.BOLD, color="#f44747")], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
                ft.Container(content=ft.Column([ft.Text("Saldo", size=12, color=COR_SUBTEXTO), ft.Text(f"R$ {resumo['saldo']:,.2f}", size=20, weight=ft.FontWeight.BOLD, color="#4ec9b0" if resumo['saldo'] >= 0 else "#f44747")], spacing=4),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16, expand=True),
            ], spacing=12)
        )
        area_resultado.controls.append(ft.Text(f"Receitas ({len(receitas)})", size=14, weight=ft.FontWeight.BOLD, color="#4ec9b0"))
        for r in receitas:
            area_resultado.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Text(r['descricao'], size=13, color=COR_TEXTO, expand=True),
                        ft.Text(r['vencimento'], size=12, color=COR_SUBTEXTO),
                        ft.Text(f"R$ {r['valor']:,.2f}", size=13, weight=ft.FontWeight.BOLD, color="#4ec9b0"),
                        ft.Text("✓" if r['pago'] else "⏳", size=13),
                    ]),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
                    border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=8)
                )
            )
        area_resultado.controls.append(ft.Text(f"Despesas ({len(despesas)})", size=14, weight=ft.FontWeight.BOLD, color="#f44747"))
        for d in despesas:
            area_resultado.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Text(d['descricao'], size=13, color=COR_TEXTO, expand=True),
                        ft.Text(d['vencimento'], size=12, color=COR_SUBTEXTO),
                        ft.Text(f"R$ {d['valor']:,.2f}", size=13, weight=ft.FontWeight.BOLD, color="#f44747"),
                        ft.Text("✓" if d['pago'] else "⏳", size=13),
                    ]),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
                    border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=8)
                )
            )
        page.update()

    def relatorio_vendas(e):
        area_resultado.controls.clear()
        vendas = buscar_todos('''
            SELECT v.*, c.nome as cliente_nome FROM vendas v
            LEFT JOIN clientes c ON v.cliente_id = c.id
            ORDER BY v.criado_em DESC
        ''')
        total = sum(v['total'] for v in vendas)
        area_resultado.controls.append(ft.Text(f"Relatório de Vendas — {len(vendas)} vendas", size=16, weight=ft.FontWeight.BOLD, color=COR_TEXTO))
        area_resultado.controls.append(
            ft.Container(
                content=ft.Row([
                    ft.Column([ft.Text("Total de vendas", size=12, color=COR_SUBTEXTO), ft.Text(f"R$ {total:,.2f}", size=24, weight=ft.FontWeight.BOLD, color=COR_AZUL)], spacing=4, expand=True),
                    ft.Column([ft.Text("Quantidade", size=12, color=COR_SUBTEXTO), ft.Text(str(len(vendas)), size=24, weight=ft.FontWeight.BOLD, color="#4ec9b0")], spacing=4),
                ]),
                bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA), border_radius=10, padding=16
            )
        )
        for v in vendas:
            area_resultado.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Text(v['cliente_nome'] or "Sem cliente", size=13, weight=ft.FontWeight.W_500, color=COR_TEXTO),
                            ft.Text(f"{v['forma_pagamento']} • {v['criado_em'][:10]}", size=12, color=COR_SUBTEXTO),
                        ], spacing=2, expand=True),
                        ft.Text(f"R$ {v['total']:,.2f}", size=14, weight=ft.FontWeight.BOLD, color="#4ec9b0"),
                    ]),
                    bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
                    border_radius=8, padding=ft.padding.symmetric(horizontal=16, vertical=10)
                )
            )
        page.update()

    def exportar_excel(e):
        try:
            from openpyxl import Workbook
            os.makedirs('relatorios', exist_ok=True)
            data    = datetime.now().strftime('%Y%m%d_%H%M%S')
            caminho = f'relatorios/erp_{data}.xlsx'
            wb      = Workbook()
            wb.remove(wb.active)

            ws1 = wb.create_sheet('Clientes')
            clientes = buscar_todos("SELECT nome, cpf_cnpj, telefone, email, cidade FROM clientes WHERE ativo=1")
            if clientes:
                ws1.append(list(clientes[0].keys()))
                for c in clientes:
                    ws1.append(list(c.values()))

            ws2 = wb.create_sheet('Estoque')
            produtos = buscar_todos("SELECT nome, categoria, unidade, preco_custo, preco_venda, estoque_atual, estoque_minimo FROM produtos WHERE ativo=1")
            if produtos:
                ws2.append(list(produtos[0].keys()))
                for p in produtos:
                    ws2.append(list(p.values()))

            ws3 = wb.create_sheet('Financeiro')
            financeiro = buscar_todos("SELECT tipo, descricao, valor, vencimento, pago, categoria FROM financeiro")
            if financeiro:
                ws3.append(list(financeiro[0].keys()))
                for f in financeiro:
                    ws3.append(list(f.values()))

            ws4 = wb.create_sheet('Vendas')
            vendas = buscar_todos("SELECT id, total, desconto, forma_pagamento, status, criado_em FROM vendas")
            if vendas:
                ws4.append(list(vendas[0].keys()))
                for v in vendas:
                    ws4.append(list(v.values()))

            wb.save(caminho)
            snack(page, f"✅ Excel exportado: {caminho}")
            os.startfile(os.path.abspath(caminho))

        except Exception as ex:
            snack(page, f"❌ Erro: {str(ex)}", "#f44747")

    def btn_relatorio(texto, icone, cor, acao):
        return ft.Container(
            content=ft.Column([
                ft.Icon(icone, color=cor, size=32),
                ft.Text(texto, size=13, color=COR_TEXTO, text_align=ft.TextAlign.CENTER),
            ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=COR_CARD, border=ft.border.all(1, COR_BORDA),
            border_radius=12, padding=20, expand=True,
            on_click=acao, ink=True
        )

    botoes = ft.Row([
        btn_relatorio("Clientes",        ft.Icons.PEOPLE,          "#4a7cf7", relatorio_clientes),
        btn_relatorio("Estoque",         ft.Icons.INVENTORY_2,     "#4ec9b0", relatorio_estoque),
        btn_relatorio("Financeiro",      ft.Icons.ACCOUNT_BALANCE, "#dcdcaa", relatorio_financeiro),
        btn_relatorio("Vendas",          ft.Icons.POINT_OF_SALE,   "#c586c0", relatorio_vendas),
        btn_relatorio("Exportar\nExcel", ft.Icons.DOWNLOAD,        "#4ec9b0", exportar_excel),
    ], spacing=12)

    return ft.Column([
        ft.Text("Relatórios", size=22, weight=ft.FontWeight.BOLD, color=COR_TEXTO),
        ft.Text("Clique em um relatório para visualizar", size=13, color=COR_SUBTEXTO),
        ft.Divider(color=COR_BORDA),
        botoes,
        ft.Container(height=10),
        area_resultado,
    ], spacing=8, expand=True)
